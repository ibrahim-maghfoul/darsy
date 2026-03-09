"""
tawjihnet_full_scraper.py
=========================
Scrapes ALL pages of https://www.tawjihnet.net/category/etudiant/
(24 pages × ~20 articles = ~480 articles)

For each article it extracts:
  - Title, date, URL, image alt
  - All paragraphs (full text)
  - All images (alt + src URL)
  - All hyperlinks (text + URL)
  - All ATTACHMENT links (PDFs, Google Drive, external docs)
  - Related articles

Outputs:
  tawjihnet_full.xlsx      — 6-sheet formatted workbook
  tawjihnet_full.json      — complete structured data
  tawjihnet_cards.csv      — flat card listing
  tawjihnet_attachments.csv — all attachment/download links only

Requirements:
  pip install requests beautifulsoup4 openpyxl

Usage:
  python tawjihnet_full_scraper.py
"""

import requests, json, csv, time, re, sys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime
from urllib.parse import urljoin

BACKEND_API = "http://localhost:5000"  # Darsy backend for incremental check

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
BASE_URL     = "https://www.tawjihnet.net"
CATEGORIES = {
    "Bac": "https://www.tawjihnet.net/category/bac/",
    "Etudiant": "https://www.tawjihnet.net/category/etudiant/",
    "College": "https://www.tawjihnet.net/category/3college-tc-bac/"
}
DELAY        = 1.5   # seconds between requests — be polite to the server
MAX_PAGES    = 1    # known total pages (update if more are added)
HEADERS      = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,ar;q=0.8,en;q=0.7",
}

# Domains/patterns to identify attachment/download links
ATTACHMENT_PATTERNS = [
    r"\.pdf($|\?)",
    r"drive\.google\.com/file/",
    r"docs\.google\.com/",
    r"\.docx?($|\?)",
    r"\.xlsx?($|\?)",
    r"\.pptx?($|\?)",
    r"\.zip($|\?)",
    r"1drv\.ms/",          # OneDrive
    r"dropbox\.com/s/",    # Dropbox
    r"mediafire\.com/",    # MediaFire
    r"4shared\.com/",      # 4shared
    r"mega\.nz/",          # MEGA
    r"archive\.org/download/",
]
ATTACHMENT_RE = re.compile("|".join(ATTACHMENT_PATTERNS), re.I)

SOCIAL_SKIP = [
    "facebook.com", "instagram.com", "twitter.com", "youtube.com",
    "whatsapp.com", "api.whatsapp.com", "t.me", "telegram.me",
    "linkedin.com",
]
NAV_SKIP = {
    "الرئيسية", "منتدى توجيه نت", "باك", "طلبة", "أقل من باك",
    "من نحن", "اتصل بنا", "تحميل المزيد من المواضيع",
    "WhatsApp", "Facebook", "Instagram", "Telegram",
    "Twitter", "YouTube", "Politique de confidentialité",
}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def fetch(url: str) -> BeautifulSoup | None:
    """Fetch URL and return BeautifulSoup, or None on failure."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
        r.encoding = "utf-8"
        return BeautifulSoup(r.text, "html.parser")
    except requests.RequestException as e:
        print(f"  [!] Failed: {url}  ({e})", file=sys.stderr)
        return None


def is_social(url: str) -> bool:
    return any(s in url for s in SOCIAL_SKIP)

def decode_cloudflare_email(cfemail: str) -> str:
    """Decode Cloudflare obfuscated email."""
    try:
        r = int(cfemail[:2], 16)
        email = ''.join([chr(int(cfemail[i:i+2], 16) ^ r) for i in range(2, len(cfemail), 2)])
        return email
    except Exception:
        return "[email protected]"


def is_attachment(url: str) -> bool:
    return bool(ATTACHMENT_RE.search(url))


def extract_deadline(text: str) -> str:
    """Pull first DD-MM-YYYY from text."""
    m = re.search(r"\b(\d{2}-\d{2}-\d{4})\b", text)
    return m.group(1) if m else ""


def classify_type(title: str) -> str:
    t = title.lower()
    if any(k in t for k in ["résultat", "resultats", "نتائج", "listes oral",
                              "listes définitives", "affectation", "classement"]):
        return "Résultats"
    if any(k in t for k in ["inscription", "concours", "مباراة", "مباريات", "wolouj"]):
        return "Concours / Inscription"
    if any(k in t for k in ["présélection", "preselection", "انتقاء", "listes convoq"]):
        return "Présélection"
    if any(k in t for k in ["دليل", "guide", "dalil"]):
        return "Guide"
    if any(k in t for k in ["تسجيل", "طلب"]):
        return "Procédure"
    return "Information"


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPING — CATEGORY PAGES
# ─────────────────────────────────────────────────────────────────────────────

def scrape_category_page(cat_name: str, cat_url: str, page_num: int) -> tuple[list[dict], bool]:
    """
    Scrape one category listing page.
    Returns (cards, has_next_page).
    """
    url = cat_url if page_num == 1 else f"{cat_url}page/{page_num}/"
    print(f"  Scraping [{cat_name}] page {page_num}: {url}")
    soup = fetch(url)
    if not soup:
        return [], False

    cards = []
    # Articles appear as h2 > a inside the main content
    for h2 in soup.select("h2 a[href]"):
        href = h2.get("href", "").strip()
        title = h2.get_text(strip=True)
        if not title or not href or BASE_URL not in href:
            continue
        # Skip navigation / category links
        if "/category/" in href or "/vb/" in href:
            continue

        # Find image near this heading
        img = h2.find_previous("img") or h2.find_next("img")
        img_alt = img.get("alt", "").strip() if img else ""
        img_src = ""
        if img:
            # Prioritize lazy load attributes over the src placeholder 
            raw_src = (img.get("data-src", "") or img.get("data-lazy-src", "") or img.get("src", "")).strip()
            if raw_src:
                if raw_src.startswith("/"):
                    raw_src = urljoin(BASE_URL, raw_src)
                # Remove WordPress thumbnail dimensions, e.g. "image-375x195.jpg" -> "image.jpg"
                img_src = re.sub(r'-\d+x\d+(?=\.(jpg|jpeg|png|gif|webp)$)', '', raw_src, flags=re.I)

        # Date: next sibling text that matches date pattern
        parent = h2.find_parent()
        date_str = ""
        if parent:
            raw = parent.get_text(" ", strip=True)
            m = re.search(r"\b(\d{2}-\d{2}-\d{4})\b", raw)
            date_str = m.group(1) if m else ""

        cards.append({
            "title":         title,
            "url":           href,
            "card_date":     date_str,
            "card_img_alt":  img_alt,
            "card_img_src":  img_src,
            "category":      cat_name,
        })

    # Check for next page
    has_next = bool(soup.find("a", string=re.compile("تحميل المزيد|page suivante|next", re.I)))
    # Also check for explicit page link
    if not has_next:
        next_link = soup.find("a", href=re.compile(rf"/page/{page_num+1}/"))
        has_next = bool(next_link)

    return cards, has_next


def collect_all_cards(max_pages: int = MAX_PAGES) -> list[dict]:
    """Collect all article cards from all category pages."""
    all_cards = []
    seen_urls = set()

    for cat_name, cat_url in CATEGORIES.items():
        print(f"\n  --- Category: {cat_name} ---")
        for page in range(1, max_pages + 1):
            cards, has_next = scrape_category_page(cat_name, cat_url, page)
            for c in cards:
                if c["url"] not in seen_urls:
                    seen_urls.add(c["url"])
                    all_cards.append(c)
            time.sleep(DELAY)
            if not has_next and page > 1:
                print(f"  → No next page found for {cat_name} after page {page}. Stopping.")
                break

    return all_cards


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPING — DETAIL PAGES
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# YOUTUBE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

YOUTUBE_RE = re.compile(
    r"(?:youtube\.com/(?:embed/|watch\?v=|shorts/)|youtu\.be/)"
    r"([a-zA-Z0-9_-]{11})",
    re.I,
)

def _extract_youtube_id(url: str) -> str | None:
    """Return the 11-char YouTube video ID from any YouTube URL/embed, or None."""
    m = YOUTUBE_RE.search(url or "")
    return m.group(1) if m else None


def _get_styles(tag) -> dict:
    """Extract color, weight, alignment, and size from a tag's style and name."""
    styles = {}
    
    # Check for bold tags
    if tag.name in ["strong", "b", "h1", "h2", "h3", "h4", "h5", "h6"]:
        styles["is_bold"] = True
    
    # Check for inline styles
    style_attr = tag.get("style", "").lower()
    
    # Color
    raw_color = ""
    color_match = re.search(r"color:\s*(#[0-9a-fA-F]{3,6}|rgb\([^)]+\)|[a-zA-Z]+)", style_attr)
    if color_match:
        raw_color = color_match.group(1).strip()
    
    # Text Alignment
    align_match = re.search(r"text-align:\s*([a-z]+)", style_attr)
    if align_match:
        styles["align"] = align_match.group(1).strip()
    elif tag.get("align"):
        styles["align"] = tag.get("align").lower()

    # Font Size
    size_match = re.search(r"font-size:\s*([^;]+)", style_attr)
    if size_match:
        styles["font_size"] = size_match.group(1).strip()

    # Font Weight
    if "font-weight" in style_attr:
        weight_match = re.search(r"font-weight:\s*([^;]+)", style_attr)
        if weight_match:
            styles["is_bold"] = "bold" in weight_match.group(1) or any(d in weight_match.group(1) for d in "6789")

    # Check for <font color="..."> tags
    font_tag = tag.find("font", color=True)
    if font_tag:
        raw_color = font_tag["color"].strip()

    if raw_color:
        c = raw_color.lower().replace(" ", "")
        if c not in ["#000", "#000000", "black", "rgb(0,0,0)", "#fff", "#ffffff", "white", "rgb(255,255,255)", "transparent", "inherit"]:
            styles["color"] = "#3aaa6a"
        else:
            styles["color"] = raw_color

    return styles

def _process_content(tag) -> str:
    """Extract text from a tag, preserving emails and line breaks."""
    # Handle Cloudflare email protection
    for cf in tag.select(".__cf_email__"):
        cf_data = cf.get("data-cfemail", "")
        if cf_data:
            cf.replace_with(decode_cloudflare_email(cf_data))
    
    # Handle line breaks explicitly to ensure second part starts at beginning of line
    for br in tag.find_all("br"):
        br.replace_with("\n")
        
    return tag.get_text(strip=True)

def _extract_list(list_tag) -> dict:
    """Extract <ul> or <ol> into a structured dict with styles."""
    items = []
    for li in list_tag.find_all("li", recursive=False):
        item_text = _process_content(li)
        item_styles = _get_styles(li)
        
        # Also check inner spans for styles if the li itself is generic
        for span in li.find_all("span", style=True):
            s = _get_styles(span)
            if s: item_styles.update(s)

        items.append({
            "text": item_text,
            "style": item_styles
        })

    return {
        "type": "list",
        "subtype": list_tag.name,
        "items": items,
        "style": _get_styles(list_tag)
    }

def _extract_table(table_tag) -> dict:
    """Extract <table> into a structured list of rows with styles."""
    rows = []
    for tr in table_tag.find_all("tr"):
        row_data = []
        for td in tr.find_all(["td", "th"]):
            cell_link = None
            # Find all links in cell and prioritize download ones
            all_a = td.find_all("a")
            if all_a:
                # Try to find one with download text
                for a in all_a:
                    a_text = a.get_text(strip=True).lower()
                    if any(kw in a_text for kw in ["تحميل", "download", "télécharger", "avis", "annonce"]):
                        cell_link = {"text": a.get_text(strip=True), "url": urljoin(BASE_URL, a.get("href", ""))}
                        break
                # Fallback to first link if no download keyword found
                if not cell_link and all_a[0].get("href"):
                    cell_link = {"text": all_a[0].get_text(strip=True), "url": urljoin(BASE_URL, all_a[0].get("href", ""))}
            
            row_data.append({
                "text": _process_content(td),
                "is_header": td.name == "th",
                "style": _get_styles(td),
                "link": cell_link
            })
        if row_data:
            rows.append(row_data)
            
    return {
        "type": "table",
        "rows": rows,
        "style": _get_styles(table_tag)
    }

# ─────────────────────────────────────────────────────────────────────────────

def scrape_detail_page(url: str) -> dict:
    """
    Scrape a single article detail page.
    Returns dict with: paragraphs, images, links, attachments, related
    """
    soup = fetch(url)
    if not soup:
        return {"paragraphs": [], "images": [], "links": [],
                "attachments": [], "related": []}

    # ── Find main article body ────────────────────────────────────────────────
    content = (
        soup.find("div", class_=re.compile(r"entry-content|post-content|article-body|article-content", re.I))
        or soup.find("article")
        or soup.find("div", class_=re.compile(r"content", re.I))
        or soup.find("main")
    )
    if not content:
        content = soup.body

    paragraphs, content_blocks, images, links, attachments, related = [], [], [], [], [], []

    # ── Sequential Content Blocks ─────────────────────────────────────────────
    # We iterate over tags that could be content blocks in their appearance order
    seen_texts = set()
    # Broaden the search to include wrappers that might contain important content or images
    all_tags = content.find_all(["p", "h3", "h4", "h5", "h6", "ul", "ol", "table", "img", "iframe", "div", "figure"])
    processed_tags = set()

    for tag in all_tags:
        if tag in processed_tags:
            continue
            
        if tag.name in ["ul", "ol"]:
            block = _extract_list(tag)
            content_blocks.append(block)
            paragraphs.extend([item["text"] for item in block["items"]])
            for d in tag.find_all(True): processed_tags.add(d)
        elif tag.name == "table":
            block = _extract_table(tag)
            content_blocks.append(block)
            for row in block["rows"]:
                paragraphs.append(" | ".join([cell["text"] for cell in row]))
            for d in tag.find_all(True): processed_tags.add(d)
        elif tag.name == "iframe":
            # ── YouTube iframe ────────────────────────────────────────────────
            iframe_src = tag.get("src", "") or tag.get("data-src", "")
            vid_id = _extract_youtube_id(iframe_src)
            if vid_id:
                content_blocks.append({
                    "type": "video",
                    "platform": "youtube",
                    "video_id": vid_id,
                    "embed_url": f"https://www.youtube.com/embed/{vid_id}",
                })
        elif tag.name == "img" or (tag.name in ["div", "figure"] and (tag.find("img") or tag.find("iframe"))):
            # ── YouTube inside a wrapper div/figure ───────────────────────────
            yt_iframe = tag.find("iframe")
            if yt_iframe and tag.name != "iframe":
                iframe_src = yt_iframe.get("src", "") or yt_iframe.get("data-src", "")
                vid_id = _extract_youtube_id(iframe_src)
                if vid_id:
                    content_blocks.append({
                        "type": "video",
                        "platform": "youtube",
                        "video_id": vid_id,
                        "embed_url": f"https://www.youtube.com/embed/{vid_id}",
                    })
                    for d in tag.find_all(True): processed_tags.add(d)
                    continue
            # ── Regular image ─────────────────────────────────────────────────
            # If it's a wrapper, we look for the img inside
            target_img = tag if tag.name == "img" else tag.find("img")
            if target_img:
                # Prioritize lazy-loaded sources
                src = (
                    target_img.get("data-src", "") or 
                    target_img.get("data-lazy-src", "") or 
                    target_img.get("src", "")
                )
                if src and not src.startswith("data:"):
                    full_src = urljoin(BASE_URL, src) if src.startswith("/") else src
                    content_blocks.append({
                        "type": "image",
                        "src": full_src,
                        "alt": target_img.get("alt", "").strip() or tag.get("title", "").strip()
                    })
                    # Mark all children (if wrapper) as processed
                    if tag.name != "img":
                        for d in tag.find_all(True): processed_tags.add(d)
        elif tag.name in ["p", "h1", "h2", "h3", "h4", "h5", "h6"]:
            # Handle standard text tags
            if len(tag.find_all("a")) == len(tag.find_all()) and tag.find_all():
                # If the entire tag is links, extract each as a block
                for a in tag.find_all("a"):
                    text = a.get_text(strip=True)
                    url = a.get("href", "")
                    if url and text:
                        full_url = urljoin(BASE_URL, url) if url.startswith("/") else url
                        content_blocks.append({
                            "type": "link",
                            "text": text,
                            "url": full_url,
                            "style": _get_styles(a)
                        })
                continue
            
            # Check for embedded download links in text
            download_links = []
            for a in tag.find_all("a"):
                a_text = a.get_text(strip=True)
                # Broader keywords for downloads
                if any(kw in a_text.lower() for kw in ["تحميل", "télécharger", "download", "notice", "guide", "dalil", "تحميل", "l'annonce", "avis"]):
                    url = a.get("href", "")
                    if url:
                        full_url = urljoin(BASE_URL, url) if url.startswith("/") else url
                        download_links.append({"text": a_text, "url": full_url, "style": _get_styles(a)})
                        # We don't remove it from the tag yet, so the text can still be processed
            
            text = _process_content(tag)
            if text and len(text) > 5 and text not in seen_texts:
                if any(skip in text for skip in ["Tawjihnet", "تابعونا", "Politique"]): continue
                seen_texts.add(text)
                paragraphs.append(text)
                content_blocks.append({
                    "type": "text",
                    "subtype": tag.name,
                    "text": text,
                    "style": _get_styles(tag)
                })
                
                # If we found download links in this paragraph, append them as blocks immediately after
                for dl in download_links:
                    content_blocks.append({
                        "type": "link",
                        "text": dl["text"],
                        "url": dl["url"],
                        "style": dl["style"]
                    })
        
        processed_tags.add(tag)

    # Fallback: if content_blocks is empty (maybe structure is deeper), 
    # use original deep search strategy for paragraphs
    if not content_blocks:
        for tag in content.find_all(["p", "li", "h3", "h4", "h5", "h6", "td"]):
            if len(tag.find_all("a")) == len(tag.find_all()) and tag.find_all(): continue
            text = _process_content(tag)
            if text and len(text) > 5 and text not in seen_texts:
                if any(skip in text for skip in ["Tawjihnet", "تابعونا", "Politique"]): continue
                seen_texts.add(text)
                paragraphs.append(text)
                content_blocks.append({
                    "type": "text", 
                    "subtype": tag.name, 
                    "text": text,
                    "style": _get_styles(tag)
                })

    # ── Images ────────────────────────────────────────────────────────────────
    seen_srcs = set()
    for img in content.find_all("img"):
        src = img.get("src", "") or img.get("data-src", "") or img.get("data-lazy-src", "")
        alt = img.get("alt", "").strip()
        w   = img.get("width", "")
        h   = img.get("height", "")
        # Skip base64 placeholders and tiny icons
        if src.startswith("data:") or "125x125" in src:
            continue
        if src and src not in seen_srcs:
            seen_srcs.add(src)
            full_src = urljoin(BASE_URL, src) if src.startswith("/") else src
            images.append({
                "alt":        alt,
                "src":        full_src,
                "dimensions": f"{w}×{h}" if w and h else "",
                "is_attachment": is_attachment(full_src),
            })

    # ── Links + Attachments ───────────────────────────────────────────────────
    seen_hrefs = set()
    for a in content.find_all("a", href=True):
        href  = a["href"].strip()
        label = _process_content(a)

        if not label or label in NAV_SKIP or len(label) < 2:
            if not is_attachment(href): continue
            label = label or "Download"
        if href in seen_hrefs:
            continue
        if is_social(href):
            continue
        if href.startswith("mailto:") or href.startswith("tel:"):
            continue
        # Skip email obfuscation links
        if "/cdn-cgi/l/email-protection" in href:
            continue

        seen_hrefs.add(href)
        full_href = urljoin(BASE_URL, href) if not href.startswith("http") else href

        link_obj = {
            "text": label,
            "url":  full_href,
            "type": "attachment" if is_attachment(full_href) else "link",
        }
        links.append(link_obj)
        if is_attachment(full_href):
            attachments.append({
                "label":    label,
                "url":      full_href,
                "format":   _guess_format(full_href),
            })

    # ── Related articles ──────────────────────────────────────────────────────
    rel_section = soup.find(string=re.compile(r"مواضيع مشابهة|articles? similaires?|related", re.I))
    if rel_section:
        parent = rel_section.find_parent()
        if parent:
            for sib in parent.find_next_siblings()[:12]:
                h = sib.find(["h2", "h3", "h4"])
                a = sib.find("a", href=True)
                if h and a:
                    related.append({
                        "title": h.get_text(strip=True),
                        "url":   a["href"] if a["href"].startswith("http") else BASE_URL + a["href"],
                    })

    return {
        "paragraphs":      paragraphs,
        "content_blocks":  content_blocks,
        "images":          images,
        "links":           links,
        "attachments":     attachments,
        "related":         related,
    }


def _guess_format(url: str) -> str:
    u = url.lower()
    if ".pdf" in u:           return "PDF"
    if "drive.google.com" in u: return "Google Drive"
    if "docs.google.com" in u:  return "Google Docs"
    if ".docx" in u or ".doc" in u: return "Word"
    if ".xlsx" in u or ".xls" in u: return "Excel"
    if ".pptx" in u or ".ppt" in u: return "PowerPoint"
    if ".zip" in u:               return "ZIP"
    if "mega.nz" in u:            return "MEGA"
    if "mediafire" in u:          return "MediaFire"
    if "dropbox" in u:            return "Dropbox"
    if "1drv.ms" in u:            return "OneDrive"
    return "Download"


# ─────────────────────────────────────────────────────────────────────────────
# INCREMENTAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_latest_scraped_id() -> int:
    """
    Ask the backend for the highest article id already stored in MongoDB.
    Falls back to reading the local tawjihnet_full.json if the backend is down.
    Returns 0 if nothing is stored yet (full scrape needed).
    """
    # 1. Try the backend API
    try:
        r = requests.get(f"{BACKEND_API}/api/news/latest-id", timeout=5)
        if r.status_code == 200:
            data = r.json()
            latest = int(data.get("latestId", 0))
            count  = int(data.get("count", 0))
            print(f"  📡 Backend reports {count} articles stored; latest ID = {latest}")
            return latest
    except Exception as e:
        print(f"  ⚠️  Backend unreachable ({e}). Falling back to local JSON...")

    # 2. Fall back to local JSON file
    local_json = "tawjihnet_full.json"
    try:
        with open(local_json, encoding="utf-8") as f:
            data = json.load(f)
        ids = [int(a["id"]) for a in data.get("articles", []) if str(a.get("id", "")).isdigit()]
        latest = max(ids) if ids else 0
        print(f"  📄 Local JSON has {len(ids)} articles; latest ID = {latest}")
        return latest
    except FileNotFoundError:
        print("  📄 No local JSON found. Starting fresh.")
        return 0


# ─────────────────────────────────────────────────────────────────────────────
# FULL PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def run():
    print(f"\n{'='*60}")
    print(f"  TAWJIHNET — Full Scraper (Bac, Etudiant, College)")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    # ── Step 1: Collect all cards from all pages ──────────────────────────────
    print("STEP 1: Collecting article cards from all category pages...")
    all_cards = collect_all_cards(MAX_PAGES)
    print(f"  → Collected {len(all_cards)} unique articles across all pages\n")

    # ── Incremental skip ──────────────────────────────────────────────────────
    print("STEP 1.5: Checking for already-imported articles...")
    latest_id = get_latest_scraped_id()
    if latest_id > 0:
        print(f"  ⏩ Skipping first {latest_id} articles (already in database).")
        # The cards are listed newest-first; IDs are assigned sequentially.
        # We need to assign IDs to find cards we haven't seen yet.
        # The safe approach: skip anything whose sequential position ≤ latest_id.
        new_cards = all_cards[latest_id:] if latest_id < len(all_cards) else []
        if not new_cards:
            print("  ✅ All articles are already up-to-date. Nothing to scrape!")
            return
        print(f"  🆕 {len(new_cards)} new articles to scrape.\n")
    else:
        new_cards = all_cards
        print(f"  🆕 Scraping all {len(new_cards)} articles (fresh start).\n")

    # ── Step 2: Visit each detail page ────────────────────────────────────────
    print("STEP 2: Scraping detail pages...")
    articles = []
    for idx, card in enumerate(new_cards, latest_id + 1):
        print(f"  [{idx:>3}/{latest_id + len(new_cards)}] {card['title'][:60]}...")
        detail = scrape_detail_page(card["url"])
        time.sleep(DELAY)

        article = {
            "id":           idx,
            "title":        card["title"],
            "url":          card["url"],
            "card_date":    card["card_date"],
            "card_img_alt": card.get("card_img_alt", ""),
            "imageUrl":     card.get("card_img_src", ""),
            "category":     card.get("category", "General"),
            "type":         classify_type(card["title"]),
            "deadline":     extract_deadline(card["title"] + " " + card.get("card_date", "")),
            "page_number":  ((idx - 1) // 20) + 1,
            # Detail content
            "paragraphs":     detail["paragraphs"],
            "content_blocks": detail["content_blocks"],
            "images":         detail["images"],
            "links":          detail["links"],
            "attachments":    detail["attachments"],
            "related":        detail["related"],
            # Quick counts
            "n_paragraphs": len(detail["paragraphs"]),
            "n_images":     len(detail["images"]),
            "n_links":      len(detail["links"]),
            "n_attachments":len(detail["attachments"]),
        }
        articles.append(article)

    # ── Step 3: Save outputs ──────────────────────────────────────────────────
    print(f"\nSTEP 3: Saving outputs ({len(articles)} new articles)...")
    # Merge new articles into the existing local JSON
    existing = []
    try:
        with open("tawjihnet_full.json", encoding="utf-8") as f:
            existing = json.load(f).get("articles", [])
    except FileNotFoundError:
        pass
    merged = existing + articles
    save_json(merged, "tawjihnet_full.json")
    save_cards_csv(merged, "tawjihnet_cards.csv")
    save_attachments_csv(merged, "tawjihnet_attachments.csv")
    build_excel(merged, "tawjihnet_full.xlsx")

    # ── Summary ───────────────────────────────────────────────────────────────
    total_p   = sum(a["n_paragraphs"] for a in articles)
    total_img = sum(a["n_images"] for a in articles)
    total_lnk = sum(a["n_links"] for a in articles)
    total_att = sum(a["n_attachments"] for a in articles)

    print(f"\n{'='*60}")
    print(f"  DONE — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")
    print(f"  Articles scraped : {len(articles)}")
    print(f"  Total paragraphs : {total_p}")
    print(f"  Total images     : {total_img}")
    print(f"  Total links      : {total_lnk}")
    print(f"  Total attachments: {total_att}")
    print(f"\n  Files saved:")
    print(f"    tawjihnet_full.xlsx")
    print(f"    tawjihnet_full.json")
    print(f"    tawjihnet_cards.csv")
    print(f"    tawjihnet_attachments.csv")
    print(f"{'='*60}\n")


# ─────────────────────────────────────────────────────────────────────────────
# SAVE FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def save_json(articles: list[dict], path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump({
            "scraped_at": datetime.now().isoformat(),
            "total": len(articles),
            "source": "https://www.tawjihnet.net/category/etudiant/",
            "articles": articles
        }, f, ensure_ascii=False, indent=2)
    print(f"  ✓ JSON saved    → {path}")


def save_cards_csv(articles: list[dict], path: str):
    fields = ["id", "page_number", "title", "type", "card_date", "deadline",
              "url", "n_paragraphs", "n_images", "n_links", "n_attachments"]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(articles)
    print(f"  ✓ Cards CSV     → {path}")


def save_attachments_csv(articles: list[dict], path: str):
    fields = ["article_id", "article_title", "article_date",
              "label", "format", "url"]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for a in articles:
            for att in a["attachments"]:
                w.writerow({
                    "article_id":    a["id"],
                    "article_title": a["title"],
                    "article_date":  a["card_date"],
                    "label":         att["label"],
                    "format":        att["format"],
                    "url":           att["url"],
                })
    print(f"  ✓ Attachments CSV → {path}")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

TYPE_COLORS = {
    "Concours / Inscription": "1565C0",
    "Résultats":              "2E7D32",
    "Présélection":           "00695C",
    "Guide":                  "6A1B9A",
    "Procédure":              "E65100",
    "Information":            "455A64",
}

def tc(t: str) -> str:
    return TYPE_COLORS.get(t, "455A64")

def T():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def F(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)

def HF(sz=10) -> Font:
    return Font(name="Arial", size=sz, bold=True, color="FFFFFF")

def BF(sz=9, bold=False, color="212121") -> Font:
    return Font(name="Arial", size=sz, bold=bold, color=color)

def AL(h="left", v="top", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row, col, value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c


def build_excel(articles: list[dict], path: str):
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_cards(wb, articles)
    _sheet_detail(wb, articles)
    _sheet_attachments(wb, articles)
    _sheet_links(wb, articles)
    _sheet_images(wb, articles)
    _sheet_summary(wb, articles)

    wb.save(path)
    print(f"  ✓ Excel saved   → {path}")


# ── Sheet 1: All Cards ────────────────────────────────────────────────────────
def _sheet_cards(wb, articles):
    ws = wb.create_sheet("📋 All Articles")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"TAWJIHNET.NET  |  /category/etudiant/  |  {len(articles)} Articles  |  All {MAX_PAGES} Pages  |  Scraped {datetime.now().strftime('%Y-%m-%d')}"
    c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    c.fill = F("1A237E")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[1].height = 38

    headers = ["#", "Pg", "Title", "Type", "Published", "Deadline", "Paragraphs", "Images", "Links", "URL"]
    widths  = [5,   5,    58,     24,     13,        13,        12,         9,       8,      50]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(2, col, h)
        c.font  = HF()
        c.fill  = F("283593")
        c.alignment = AL("center", "center", False)
        c.border = T()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    for i, a in enumerate(articles):
        row = i + 3
        bg  = "F0F4FF" if i % 2 == 0 else "FFFFFF"
        clr = tc(a["type"])

        for col, (val, bold, fg, ha) in enumerate([
            (a["id"],          True,  "212121", "center"),
            (a["page_number"], False, "757575", "center"),
            (a["title"],       True,  "1A237E", "left"),
            (None,             True,  "FFFFFF", "center"),   # type — colored
            (a["card_date"],   False, "424242", "center"),
            (a["deadline"] or "—", bool(a["deadline"]), "C62828" if a["deadline"] else "9E9E9E", "center"),
            (a["n_paragraphs"], False, "424242", "center"),
            (a["n_images"],    False, "424242", "center"),
            (a["n_links"],     False, "424242", "center"),
            (a["url"],         False, "1565C0", "left"),
        ], 1):
            if col == 4:
                c = ws.cell(row, col, a["type"])
                c.font  = Font(name="Arial", size=8, bold=True, color="FFFFFF")
                c.fill  = F(clr)
            else:
                c = ws.cell(row, col, val)
                c.font  = Font(name="Arial", size=9, bold=bold, color=fg)
                c.fill  = F("FFF3E0" if col == 6 and a["deadline"] else bg)
            c.alignment = AL(ha, "top", False)
            c.border = T()
        ws.row_dimensions[row].height = 22


# ── Sheet 2: Full Detail Content ──────────────────────────────────────────────
def _sheet_detail(wb, articles):
    ws = wb.create_sheet("📄 Detail Content")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "FULL DETAIL CONTENT  |  Paragraphs · Links · Images · Attachments · Related"
    c.font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill  = F("4A148C")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[1].height = 34

    headers = ["Art. #", "Title", "Content Type", "Content / Text", "Date", "URL / Source"]
    widths  = [7,        48,     20,              92,              13,     60]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(2, col, h)
        c.font = HF()
        c.fill = F("6A1B9A")
        c.alignment = AL("center", "center", False)
        c.border = T()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    CTYPE = {
        "summary":    ("EDE7F6", "6A1B9A", "📝 Summary"),
        "paragraph":  ("FAFAFA", "37474F", "📃 Paragraph"),
        "link":       ("E8F5E9", "1B5E20", "🔗 Link"),
        "attachment": ("FFF3E0", "E65100", "📎 Attachment"),
        "image":      ("FFF8E1", "BF360C", "🖼 Image"),
        "related":    ("E3F2FD", "0D47A1", "🔄 Related"),
    }

    row = 3
    for a in articles:
        clr = tc(a["type"])
        # Section header band
        ws.merge_cells(f"A{row}:F{row}")
        hdr = ws.cell(row, 1)
        hdr.value = f"  #{a['id']}  |  {a['title']}  |  {a['type']}  |  {a['card_date']}"
        hdr.font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        hdr.fill  = F(clr)
        hdr.alignment = AL("left", "center", False)
        ws.row_dimensions[row].height = 22
        row += 1

        all_rows = []
        for block in a.get("content_blocks", []):
            btype = block["type"]
            if btype == "text":
                label = "paragraph"
                if block.get("subtype","").startswith("h"): label = "header"
                all_rows.append((label, block["text"], ""))
            elif btype == "list":
                items_str = "\n".join([f"• {item['text']}" for item in block["items"]])
                all_rows.append(("list", items_str, ""))
            elif btype == "table":
                # Format table as a simple text block for the cell
                table_str = ""
                for r_idx, table_row in enumerate(block["rows"]):
                    row_str = " | ".join([cell["text"] for cell in table_row])
                    table_str += row_str + "\n"
                    if r_idx == 0: # Header separator
                        table_str += "-" * min(len(row_str), 50) + "\n"
                all_rows.append(("table", table_str.strip(), ""))

        for lk in a["links"]:
            ctype = "attachment" if lk.get("type") == "attachment" else "link"
            all_rows.append((ctype, lk["text"], lk["url"]))
        for img in a["images"]:
            label = img.get("alt","") or img.get("src","")
            all_rows.append(("image", f'{label}  [{img.get("dimensions","")}]', img.get("src","")))
        for rel in a["related"]:
            all_rows.append(("related", rel.get("title",""), rel.get("url","")))

        # Update CTYPE mapping to include new types
        CTYPE.update({
            "header":     ("E1BEE7", "4A148C", "🏷 Header"),
            "list":       ("F3E5F5", "4A148C", "📝 List"),
            "table":      ("E0F7FA", "006064", "📊 Table"),
        })

        for ctype, text, url in all_rows:
            bg_c, fg_c, label = CTYPE.get(ctype, ("FFFFFF", "212121", ctype))
            for col, val in enumerate(
                [a["id"], a["title"], label, text, a["card_date"], url or a["url"]], 1
            ):
                c = ws.cell(row, col, val)
                c.font = Font(name="Arial", size=9,
                              bold=(ctype == "summary" and col == 4),
                              color=fg_c if col == 4 else "424242")
                c.fill = F(bg_c)
                c.alignment = AL("left", "top")
                c.border = T()
            ws.row_dimensions[row].height = max(18, min(72, len(str(text)) // 7 + 15))
            row += 1
        row += 1  # blank gap between articles


# ── Sheet 3: Attachments ──────────────────────────────────────────────────────
def _sheet_attachments(wb, articles):
    ws = wb.create_sheet("📎 Attachments")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "ALL ATTACHMENT & DOWNLOAD LINKS  |  PDFs · Google Drive · Docs · Archives"
    c.font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill  = F("BF360C")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[1].height = 32

    headers = ["Art. #", "Article Title", "Label / Description", "Format", "Download URL"]
    widths  = [7,        50,              50,                    14,       70]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(2, col, h)
        c.font = HF()
        c.fill = F("D84315")
        c.alignment = AL("center", "center", False)
        c.border = T()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    FORMAT_COLORS = {
        "PDF": "B71C1C", "Google Drive": "1565C0", "Google Docs": "1B5E20",
        "Word": "0D47A1", "Excel": "1B5E20", "PowerPoint": "E65100",
        "ZIP": "4A148C", "MEGA": "880E4F", "MediaFire": "37474F",
        "Dropbox": "006064", "OneDrive": "0277BD", "Download": "455A64",
    }

    row = 3
    idx = 0
    for a in articles:
        for att in a["attachments"]:
            bg = "FFF3E0" if idx % 2 == 0 else "FFFFFF"
            fmt_color = FORMAT_COLORS.get(att["format"], "455A64")

            ws.cell(row, 1, a["id"]).font = BF(bold=True)
            ws.cell(row, 1).fill = F(bg)
            ws.cell(row, 1).alignment = AL("center", "top", False)
            ws.cell(row, 1).border = T()

            ws.cell(row, 2, a["title"]).font = BF(color="1A237E")
            ws.cell(row, 2).fill = F(bg)
            ws.cell(row, 2).alignment = AL("left", "top")
            ws.cell(row, 2).border = T()

            ws.cell(row, 3, att["label"]).font = BF()
            ws.cell(row, 3).fill = F(bg)
            ws.cell(row, 3).alignment = AL("left", "top")
            ws.cell(row, 3).border = T()

            c4 = ws.cell(row, 4, att["format"])
            c4.font = Font(name="Arial", size=8, bold=True, color="FFFFFF")
            c4.fill = F(fmt_color)
            c4.alignment = AL("center", "top", False)
            c4.border = T()

            ws.cell(row, 5, att["url"]).font = BF(color="1565C0")
            ws.cell(row, 5).fill = F(bg)
            ws.cell(row, 5).alignment = AL("left", "top")
            ws.cell(row, 5).border = T()

            ws.row_dimensions[row].height = 20
            row += 1
            idx += 1

    if idx == 0:
        ws.merge_cells(f"A3:E3")
        ws.cell(3, 1, "No attachments found — run scraper on live site to collect real data.")
        ws.cell(3, 1).font = BF(color="9E9E9E")
        ws.cell(3, 1).alignment = AL("center", "center", False)


# ── Sheet 4: All Links ────────────────────────────────────────────────────────
def _sheet_links(wb, articles):
    ws = wb.create_sheet("🔗 All Links")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "ALL HYPERLINKS  |  Internal + External + Downloads"
    c.font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill  = F("00695C")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[1].height = 30

    for col, (h, w) in enumerate(zip(
        ["Art.#", "Article Title", "Link Text", "Link Type", "URL"],
        [7, 50, 44, 14, 68]
    ), 1):
        c = ws.cell(2, col, h)
        c.font = HF()
        c.fill = F("00796B")
        c.alignment = AL("center", "center", False)
        c.border = T()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    row = 3
    for a in articles:
        # Card link itself
        for col, (val, fg) in enumerate([
            (a["id"], "212121"), (a["title"], "004D40"),
            ("→ Article page", "004D40"), ("Article", "FFFFFF"), (a["url"], "004D40")
        ], 1):
            c = ws.cell(row, col, val)
            c.font = Font(name="Arial", size=9, bold=True, color=fg)
            c.fill = F("E0F2F1")
            c.alignment = AL("left", "center", False)
            c.border = T()
        ws.cell(row, 4).fill = F("00695C")
        ws.row_dimensions[row].height = 17
        row += 1
        # Detail links
        for lk in a["links"]:
            bg = "FFF3E0" if lk.get("type") == "attachment" else ("F9FBE7" if row % 2 == 0 else "FFFFFF")
            ltype = lk.get("type", "link").title()
            for col, (val, fg) in enumerate([
                (a["id"], "424242"), (a["title"], "424242"),
                (lk["text"], "37474F"), (ltype, "FFFFFF"), (lk["url"], "0D47A1")
            ], 1):
                c = ws.cell(row, col, val)
                c.font = Font(name="Arial", size=9, color=fg)
                c.fill = F(bg)
                c.alignment = AL("left", "center", False)
                c.border = T()
            ws.cell(row, 4).font = Font(name="Arial", size=8, bold=True, color="FFFFFF")
            tc_lnk = "BF360C" if lk.get("type") == "attachment" else "0D47A1"
            ws.cell(row, 4).fill = F(tc_lnk)
            ws.row_dimensions[row].height = 17
            row += 1


# ── Sheet 5: Images ───────────────────────────────────────────────────────────
def _sheet_images(wb, articles):
    ws = wb.create_sheet("🖼 Images")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "ALL IMAGES  |  Alt Text · Source URL · Dimensions"
    c.font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill  = F("E65100")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[1].height = 30

    for col, (h, w) in enumerate(zip(
        ["Art.#", "Article Title", "Alt Text", "Dimensions", "Image URL"],
        [7, 50, 50, 13, 70]
    ), 1):
        c = ws.cell(2, col, h)
        c.font = HF()
        c.fill = F("BF360C")
        c.alignment = AL("center", "center", False)
        c.border = T()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    row = 3
    for a in articles:
        for img in a["images"]:
            bg = "FFF8E1" if row % 2 == 0 else "FFFDE7"
            for col, val in enumerate([
                a["id"], a["title"], img.get("alt",""), img.get("dimensions",""), img.get("src","")
            ], 1):
                c = ws.cell(row, col, val)
                c.font = Font(name="Arial", size=9,
                              color="E65100" if col == 5 else "424242")
                c.fill = F(bg)
                c.alignment = AL("left", "top")
                c.border = T()
            ws.row_dimensions[row].height = 18
            row += 1


# ── Sheet 6: Summary ──────────────────────────────────────────────────────────
def _sheet_summary(wb, articles):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"SCRAPE SUMMARY  |  {len(articles)} Articles  |  {datetime.now().strftime('%Y-%m-%d')}"
    c.font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill  = F("263238")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[1].height = 30

    # Global stats
    stats = [
        ("Total Articles",    len(articles)),
        ("Total Paragraphs",  sum(a["n_paragraphs"] for a in articles)),
        ("Total Images",      sum(a["n_images"] for a in articles)),
        ("Total Links",       sum(a["n_links"] for a in articles)),
        ("Total Attachments", sum(a["n_attachments"] for a in articles)),
        ("Articles w/ Deadline", sum(1 for a in articles if a["deadline"])),
        ("Pages Scraped",     MAX_PAGES),
    ]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 12

    # Stats section header
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = "Global Statistics"
    c.font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill  = F("37474F")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[2].height = 22

    for i, (label, val) in enumerate(stats):
        bg = "ECEFF1" if i % 2 == 0 else "FFFFFF"
        c = ws.cell(i+3, 1, label)
        c.font = BF(bold=True, color="37474F")
        c.fill = F(bg)
        c.alignment = AL("left", "center", False)
        c.border = T()
        c2 = ws.cell(i+3, 2, val)
        c2.font = Font(name="Arial", size=11, bold=True, color="1A237E")
        c2.fill = F(bg)
        c2.alignment = AL("center", "center", False)
        c2.border = T()
        ws.row_dimensions[i+3].height = 24

    # Type breakdown
    start_row = len(stats) + 5
    ws.merge_cells(f"A{start_row}:D{start_row}")
    c = ws.cell(start_row, 1, "Breakdown by Article Type")
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = F("37474F")
    c.alignment = AL("center", "center", False)
    ws.row_dimensions[start_row].height = 22

    for col, (h, w) in enumerate(zip(["Type","Count","% of Total","Avg. Attachments"], [26,10,14,20]), 1):
        c = ws.cell(start_row+1, col, h)
        c.font = HF()
        c.fill = F("546E7A")
        c.alignment = AL("center","center",False)
        c.border = T()

    type_arts = Counter(a["type"] for a in articles)
    total = len(articles)
    for i, (t, cnt) in enumerate(sorted(type_arts.items(), key=lambda x: -x[1])):
        bg = "ECEFF1" if i % 2 == 0 else "FFFFFF"
        avg_att = sum(a["n_attachments"] for a in articles if a["type"] == t) / max(cnt, 1)
        clr = tc(t)
        r = start_row + 2 + i
        c1 = ws.cell(r, 1, t)
        c1.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c1.fill = F(clr)
        c1.alignment = AL("center","center",False)
        c1.border = T()
        for col, val in [(2, cnt), (3, f"{cnt/total*100:.1f}%"), (4, f"{avg_att:.1f}")]:
            c = ws.cell(r, col, val)
            c.font = BF()
            c.fill = F(bg)
            c.alignment = AL("center","center",False)
            c.border = T()
        ws.row_dimensions[r].height = 22


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    run()